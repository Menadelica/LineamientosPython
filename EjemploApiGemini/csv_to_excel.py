#!/usr/bin/env python3
"""
Script para convertir los resultados CSV a Excel con formato mejorado
"""

import pandas as pd
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_excel_report():
    """Crea un reporte en Excel con los resultados de la automatización"""
    
    # Rutas de archivos
    csv_file = Path("data/output/results.csv")
    report_file = Path("data/output/execution_report.json")
    excel_file = Path("data/output/resultados_completos.xlsx")
    
    # Verificar que existan los archivos
    if not csv_file.exists():
        print(f"❌ No se encontró el archivo: {csv_file}")
        return
    
    if not report_file.exists():
        print(f"❌ No se encontró el archivo: {report_file}")
        return
    
    print("📊 Creando reporte en Excel...")
    
    # Leer datos
    df_results = pd.read_csv(csv_file)
    with open(report_file, 'r', encoding='utf-8') as f:
        execution_data = json.load(f)
    
    # Crear el workbook de Excel
    wb = Workbook()
    
    # === HOJA 1: RESUMEN EJECUTIVO ===
    ws_summary = wb.active
    ws_summary.title = "Resumen Ejecutivo"
    
    # Configurar estilos
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    subheader_font = Font(bold=True, size=12)
    data_font = Font(size=11)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título principal
    ws_summary['A1'] = "🤖 REPORTE DE AUTOMATIZACIÓN GEMINI"
    ws_summary['A1'].font = Font(bold=True, size=16, color="366092")
    ws_summary.merge_cells('A1:E1')
    
    # Información del reporte
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_summary['A3'] = f"📅 Generado: {current_time}"
    ws_summary['A3'].font = data_font
    
    # Estadísticas de ejecución
    summary = execution_data['execution_summary']
    
    ws_summary['A5'] = "📊 ESTADÍSTICAS DE EJECUCIÓN"
    ws_summary['A5'].font = subheader_font
    ws_summary['A5'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws_summary.merge_cells('A5:E5')
    
    stats_data = [
        ("Tiempo de Inicio", summary['start_time']),
        ("Tiempo de Fin", summary['end_time']),
        ("Duración (segundos)", f"{summary['duration_seconds']:.2f}"),
        ("Total de Elementos", summary['total_items']),
        ("Elementos Exitosos", summary['successful_items']),
        ("Elementos Fallidos", summary['failed_items']),
        ("Tasa de Éxito", f"{summary['success_rate_percent']}%")
    ]
    
    for i, (label, value) in enumerate(stats_data, start=6):
        ws_summary[f'A{i}'] = label
        ws_summary[f'B{i}'] = value
        ws_summary[f'A{i}'].font = subheader_font
        ws_summary[f'B{i}'].font = data_font
        
        # Color de fondo alternado
        if i % 2 == 0:
            ws_summary[f'A{i}'].fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
            ws_summary[f'B{i}'].fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
    
    # Resumen de resultados
    ws_summary['A15'] = "🎯 RESUMEN DE PROMPTS PROCESADOS"
    ws_summary['A15'].font = subheader_font
    ws_summary['A15'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws_summary.merge_cells('A15:E15')
    
    for i, row in enumerate(df_results.itertuples(), start=16):
        ws_summary[f'A{i}'] = f"ID {row.id}"
        ws_summary[f'B{i}'] = row.original_prompt[:80] + "..." if len(row.original_prompt) > 80 else row.original_prompt
        ws_summary[f'C{i}'] = f"{row.response_length} caracteres"
        ws_summary[f'D{i}'] = row.status.upper()
        
        ws_summary[f'A{i}'].font = subheader_font
        ws_summary[f'B{i}'].font = data_font
        ws_summary[f'C{i}'].font = data_font
        ws_summary[f'D{i}'].font = Font(bold=True, color="008000" if row.status == "completed" else "FF0000")
    
    # Ajustar anchos de columna
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 60
    ws_summary.column_dimensions['C'].width = 20
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 15
    
    # === HOJA 2: RESULTADOS DETALLADOS ===
    ws_details = wb.create_sheet(title="Resultados Detallados")
    
    # Encabezados
    headers = ["ID", "Prompt Original", "Respuesta Generada", "Estado", "Modelo Usado", "Longitud Respuesta", "Procesado"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws_details.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row_idx, row in enumerate(df_results.itertuples(), start=2):
        # Truncar respuesta si es muy larga para Excel
        response_text = row.generated_response
        if len(response_text) > 32767:  # Límite de Excel
            response_text = response_text[:32760] + "..."
        
        data_row = [
            row.id,
            row.original_prompt,
            response_text,
            row.status,
            row.model_used,
            row.response_length,
            row.processed_at
        ]
        
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws_details.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = border
            
            # Alineación especial para diferentes columnas
            if col_idx in [1, 5, 6]:  # ID, longitud, fecha
                cell.alignment = Alignment(horizontal='center')
            elif col_idx in [2, 3]:  # Prompt y respuesta
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Ajustar anchos y alturas
    ws_details.column_dimensions['A'].width = 8
    ws_details.column_dimensions['B'].width = 50
    ws_details.column_dimensions['C'].width = 80
    ws_details.column_dimensions['D'].width = 12
    ws_details.column_dimensions['E'].width = 20
    ws_details.column_dimensions['F'].width = 15
    ws_details.column_dimensions['G'].width = 25
    
    # Ajustar altura de filas para texto largo
    for row in range(2, len(df_results) + 2):
        ws_details.row_dimensions[row].height = 60
    
    # === HOJA 3: DATOS RAW (CSV) ===
    ws_raw = wb.create_sheet(title="Datos CSV Raw")
    
    # Copiar datos del DataFrame
    for r in dataframe_to_rows(df_results, index=False, header=True):
        ws_raw.append(r)
    
    # Formatear encabezados
    for cell in ws_raw[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Guardar archivo
    excel_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_file)
    
    print(f"✅ Reporte Excel creado exitosamente: {excel_file}")
    print(f"📊 El archivo contiene {len(df_results)} registros en {len(wb.worksheets)} hojas")
    
    return excel_file

if __name__ == "__main__":
    create_excel_report()
